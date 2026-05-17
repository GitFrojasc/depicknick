from io import BytesIO
from django.contrib import admin
from django.utils.html import format_html
from django.core.mail import EmailMultiAlternatives
from django.template.loader import render_to_string
from django.http import HttpResponse
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from import_export import resources, fields
from import_export.admin import ImportExportModelAdmin
from import_export.widgets import ForeignKeyWidget
from .models import (
    Productor, Categoria, Producto, Canasto, Pedido, ItemPedido, Suscripcion,
    Mercado, Receta, IngredienteReceta, Paquete, ItemPaquete,
    ClienteCorporativo, DocumentoRAG, ConversacionAgente, MensajeAgente,
)


class ProductoResource(resources.ModelResource):
    productor = fields.Field(
        column_name='productor',
        attribute='productor',
        widget=ForeignKeyWidget(Productor, field='nombre')
    )
    categoria = fields.Field(
        column_name='categoria_tipo',
        attribute='categoria',
        widget=ForeignKeyWidget(Categoria, field='tipo')
    )

    class Meta:
        model = Producto
        fields = (
            'id', 'nombre', 'descripcion', 'ingredientes', 'historia',
            'productor', 'categoria', 'precio', 'unidad',
            'precio_mayor', 'unidad_mayor', 'caducidad',
            'tiempo_espera', 'cobertura_envio',
            'stock', 'stock_minimo', 'destacado', 'certificado',
        )
        export_order = fields
        import_id_fields = ['nombre']

    def skip_row(self, instance, original, row, import_validation_errors=None):
        if str(row.get('nombre', '')).startswith('(ejemplo)'):
            return True
        return super().skip_row(instance, original, row, import_validation_errors)


class ItemPedidoInline(admin.TabularInline):
    model = ItemPedido
    extra = 0
    readonly_fields = ('subtotal',)


def _contexto_invitacion(productor, request):
    return {
        'productor': productor,
        'productos': Producto.objects.filter(productor=productor).select_related('categoria'),
        'request': request,
    }


def _generar_excel_onboarding(productor):
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Side, Border
    wb = openpyxl.Workbook()

    # --- Hoja 1: Instrucciones ---
    ws_info = wb.active
    ws_info.title = "Instrucciones"
    ws_info.column_dimensions['A'].width = 90

    fill_verde = PatternFill("solid", fgColor="2D5A27")
    fill_crema = PatternFill("solid", fgColor="F5F0E8")
    fill_cafe = PatternFill("solid", fgColor="5C3D2E")

    instrucciones = [
        ("dePicknick — Formulario de productos", Font(name="Arial", size=16, bold=True, color="FFFFFF"), fill_verde, 36),
        (f"Productor: {productor.nombre}  ·  {productor.ubicacion}", Font(name="Arial", size=11, color="FFFFFF", italic=True), fill_verde, 22),
        ("", None, fill_verde, 10),
        ("¿Cómo usar este archivo?", Font(name="Arial", size=12, bold=True, color="5C3D2E"), fill_crema, 22),
        ("1. Ve a la hoja «Productos» (pestaña abajo).", Font(name="Arial", size=10, color="333333"), fill_crema, 18),
        ("2. Completa cada columna para tus productos. Las columnas con * son obligatorias.", Font(name="Arial", size=10, color="333333"), fill_crema, 18),
        ("3. Agrega tantas filas como necesites. La fila gris es un ejemplo — no la borres.", Font(name="Arial", size=10, color="333333"), fill_crema, 18),
        ("4. Envía el archivo completado a hola@depicknick.com y nosotros lo subimos.", Font(name="Arial", size=10, color="333333"), fill_crema, 18),
        ("", None, fill_crema, 10),
        ("Guía de columnas:", Font(name="Arial", size=11, bold=True, color="FFFFFF"), fill_cafe, 22),
        ("nombre              → Nombre corto del producto", Font(name="Arial", size=10, color="333333"), None, 18),
        ("descripcion         → Descripción para los clientes (1-2 oraciones)", Font(name="Arial", size=10, color="333333"), None, 18),
        ("ingredientes        → Lista de ingredientes o materias primas", Font(name="Arial", size=10, color="333333"), None, 18),
        ("historia            → Historia detrás del producto (opcional pero recomendado)", Font(name="Arial", size=10, color="333333"), None, 18),
        ("precio              → Precio de venta al detal en pesos COP (ej: 28000)", Font(name="Arial", size=10, color="333333"), None, 18),
        ("unidad              → Unidad de medida (ej: 250g, kg, unidad, 500ml)", Font(name="Arial", size=10, color="333333"), None, 18),
        ("precio_mayor        → Precio al por mayor para restaurantes (opcional)", Font(name="Arial", size=10, color="333333"), None, 18),
        ("unidad_mayor        → Unidad mayoreo (ej: caja x 12, bulto x 25kg)", Font(name="Arial", size=10, color="333333"), None, 18),
        ("caducidad           → Vida útil (ej: 6 meses sellado, refrigerado 15 días)", Font(name="Arial", size=10, color="333333"), None, 18),
        ("tiempo_espera       → Días de espera para preparar el pedido (ej: 3)", Font(name="Arial", size=10, color="333333"), None, 18),
        ("cobertura_envio     → Escribe: local  o  nacional", Font(name="Arial", size=10, color="333333"), None, 18),
        ("stock               → Unidades disponibles hoy", Font(name="Arial", size=10, color="333333"), None, 18),
        ("stock_minimo        → Número mínimo antes de alertarte (recomendado: 5)", Font(name="Arial", size=10, color="333333"), None, 18),
        ("destacado           → Escribe TRUE si quieres que aparezca primero", Font(name="Arial", size=10, color="333333"), None, 18),
        ("certificado         → Escribe TRUE si tienes certificación oficial", Font(name="Arial", size=10, color="333333"), None, 18),
        ("productor           → Tu nombre exactamente como lo registraste", Font(name="Arial", size=10, color="333333"), None, 18),
        ("categoria_tipo      → Escribe una de: despensa / gourmet / ancheta / picknick", Font(name="Arial", size=10, color="333333"), None, 18),
    ]

    for texto, fuente, relleno, altura in instrucciones:
        r = ws_info.max_row + 1 if ws_info.max_row > 0 else 1
        ws_info.append([texto])
        if fuente:
            ws_info.cell(row=r, column=1).font = fuente
        if relleno:
            ws_info.cell(row=r, column=1).fill = relleno
        ws_info.cell(row=r, column=1).alignment = Alignment(vertical='center', wrap_text=True)
        ws_info.row_dimensions[r].height = altura

    # --- Hoja 2: Productos (importable) ---
    ws = wb.create_sheet("Productos")

    columnas = [
        ('nombre *', 'nombre', 28),
        ('descripcion *', 'descripcion', 38),
        ('ingredientes', 'ingredientes', 28),
        ('historia', 'historia', 35),
        ('precio *', 'precio', 16),
        ('unidad *', 'unidad', 14),
        ('precio_mayor', 'precio_mayor', 16),
        ('unidad_mayor', 'unidad_mayor', 18),
        ('caducidad', 'caducidad', 22),
        ('tiempo_espera', 'tiempo_espera', 16),
        ('cobertura_envio', 'cobertura_envio', 18),
        ('stock', 'stock', 12),
        ('stock_minimo', 'stock_minimo', 14),
        ('destacado', 'destacado', 12),
        ('certificado', 'certificado', 14),
        ('productor', 'productor', 26),
        ('categoria_tipo', 'categoria_tipo', 18),
    ]

    fill_header = PatternFill("solid", fgColor="2D5A27")
    fill_ejemplo = PatternFill("solid", fgColor="EEEEEE")
    font_header = Font(name="Arial", size=9, bold=True, color="FFFFFF")
    font_ejemplo = Font(name="Arial", size=9, italic=True, color="999999")
    font_data = Font(name="Arial", size=10)

    # Row 1: headers (field names — for import)
    for i, (label, field, width) in enumerate(columnas, 1):
        col = get_column_letter(i)
        ws.column_dimensions[col].width = width
        cell = ws.cell(row=1, column=i, value=field)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.comment = None
    ws.row_dimensions[1].height = 28

    # Row 2: example (skipped on import via skip_row)
    productos_qs = Producto.objects.filter(productor=productor).select_related('categoria')
    ejemplo = [
        '(ejemplo) Café Especial Tostado', 'Café arábica de altura, tostado artesanalmente',
        '100% café arábica', 'Crece a 1.800m s.n.m., cosecha manual',
        '28000', '250g', '50000', 'caja x 4',
        '6 meses sellado', '3', 'nacional', '30', '5', 'TRUE', 'FALSE',
        productor.nombre, 'gourmet',
    ]
    for i, val in enumerate(ejemplo, 1):
        cell = ws.cell(row=2, column=i, value=val)
        cell.font = font_ejemplo
        cell.fill = fill_ejemplo
        cell.alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[2].height = 18

    # Rows 3+: existing products pre-filled
    row_num = 3
    for p in productos_qs:
        data = [
            p.nombre, p.descripcion, p.ingredientes, p.historia,
            float(p.precio) if p.precio else '',
            p.unidad,
            float(p.precio_mayor) if p.precio_mayor else '',
            p.unidad_mayor or '',
            p.caducidad, p.tiempo_espera, p.cobertura_envio,
            p.stock, p.stock_minimo,
            'TRUE' if p.destacado else 'FALSE',
            'TRUE' if p.certificado else 'FALSE',
            p.productor.nombre if p.productor else '',
            p.categoria.tipo if p.categoria else '',
        ]
        for i, val in enumerate(data, 1):
            cell = ws.cell(row=row_num, column=i, value=val)
            cell.font = font_data
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        ws.row_dimensions[row_num].height = 18
        row_num += 1

    # 5 blank rows for new products
    for _ in range(5):
        for i in range(1, len(columnas) + 1):
            ws.cell(row=row_num, column=i, value='')
        ws.row_dimensions[row_num].height = 18
        row_num += 1

    ws.freeze_panes = 'A2'

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


@admin.register(Mercado)
class MercadoAdmin(admin.ModelAdmin):
    list_display = ('nombre', 'ciudad', 'acopiador', 'activo')
    list_filter = ('activo',)
    search_fields = ('nombre', 'ciudad')


@admin.register(Productor)
class ProductorAdmin(admin.ModelAdmin):
    list_display = ('nombre', 'mercado', 'estado_editorial', 'tipo_acuerdo', 'puntaje', 'ubicacion', 'email', 'activo')
    list_filter = ('mercado', 'estado_editorial', 'tipo_acuerdo', 'activo')
    search_fields = ('nombre', 'ubicacion', 'email')
    readonly_fields = ('puntaje',)
    actions = ['enviar_invitacion', 'previsualizar_invitacion']

    @admin.display(description='Completitud')
    def puntaje(self, obj):
        p = obj.puntaje_completitud
        color = '#38a169' if p >= 80 else '#dd6b20' if p >= 50 else '#e53e3e'
        return format_html('<span style="color:{};font-weight:bold;">{}%</span>', color, p)

    def _enviar_correo_productor(self, request, productor):
        ctx = _contexto_invitacion(productor, request)
        html = render_to_string('emails/invitacion_productor.html', ctx)
        msg = EmailMultiAlternatives(
            subject=f'Tu marca en dePicknick — {productor.nombre}',
            body=f'Hola {productor.nombre}, te presentamos tu perfil en dePicknick.',
            from_email='dePicknick <onboarding@resend.dev>',
            to=[productor.email],
        )
        msg.attach_alternative(html, 'text/html')
        excel = _generar_excel_onboarding(productor)
        msg.attach(
            f'depicknick_catalogo_{productor.nombre.replace(" ", "_")}.xlsx',
            excel,
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        msg.send(fail_silently=False)

    @admin.action(description='Enviar correo de invitación a productores seleccionados')
    def enviar_invitacion(self, request, queryset):
        enviados, sin_email, errores = 0, 0, 0
        for productor in queryset:
            if not productor.email:
                sin_email += 1
                continue
            try:
                self._enviar_correo_productor(request, productor)
                enviados += 1
            except Exception as e:
                errores += 1
                self.message_user(request, f'Error enviando a {productor.nombre}: {e}', level='error')
        if enviados:
            self.message_user(request, f'{enviados} correo(s) enviado(s) con Excel adjunto.')
        if sin_email:
            self.message_user(request, f'{sin_email} productor(es) sin email — no se les envió.', level='warning')

    def save_model(self, request, obj, form, change):
        is_new = not obj.pk
        super().save_model(request, obj, form, change)
        if is_new and obj.email:
            try:
                self._enviar_correo_productor(request, obj)
                self.message_user(request, f'Correo de bienvenida enviado a {obj.email} con el Excel adjunto.')
            except Exception as e:
                self.message_user(request, f'Productor creado pero error al enviar correo: {e}', level='warning')

    @admin.action(description='Previsualizar correo (abre en el navegador)')
    def previsualizar_invitacion(self, request, queryset):
        productor = queryset.first()
        ctx = _contexto_invitacion(productor, request)
        html = render_to_string('emails/invitacion_productor.html', ctx)
        return HttpResponse(html)


@admin.register(Categoria)
class CategoriaAdmin(admin.ModelAdmin):
    list_display = ('nombre', 'tipo', 'emoji')


@admin.register(Producto)
class ProductoAdmin(ImportExportModelAdmin):
    resource_classes = [ProductoResource]
    list_display = ('nombre', 'categoria', 'productor', 'precio', 'precio_final_display', 'unidad', 'indicador_stock', 'stock', 'estado_editorial', 'alerta_venc', 'destacado', 'disponible')
    list_filter = ('categoria', 'estado_editorial', 'disponible', 'destacado', 'certificado', 'productor')
    search_fields = ('nombre', 'descripcion')
    list_editable = ('precio', 'stock', 'destacado')
    readonly_fields = ('disponible', 'creado', 'alerta_venc', 'precio_final_display')
    fieldsets = (
        ('Información', {
            'fields': ('nombre', 'descripcion', 'historia', 'sugerencias_uso', 'productor', 'categoria', 'estado_editorial')
        }),
        ('Precio y presentación', {
            'fields': ('precio', 'unidad', 'precio_mayor', 'unidad_mayor', 'foto', 'video_url', 'destacado', 'certificado')
        }),
        ('Temporada y vencimiento', {
            'fields': ('meses_temporada', 'fecha_vencimiento', 'alerta_venc', 'precio_final_display'),
        }),
        ('Inventario', {
            'fields': ('stock', 'stock_minimo', 'disponible'),
        }),
        ('Sistema', {
            'fields': ('creado',),
            'classes': ('collapse',)
        }),
    )

    @admin.display(description='Alerta venc.')
    def alerta_venc(self, obj):
        a = obj.alerta_vencimiento
        if a == 'rojo':
            return format_html('<span style="color:#e53e3e;font-weight:bold;">🔴 Vence pronto</span>')
        if a == 'amarillo':
            return format_html('<span style="color:#dd6b20;font-weight:bold;">🟡 {} días</span>', obj.dias_para_vencer)
        return '–'

    @admin.display(description='Precio final')
    def precio_final_display(self, obj):
        pf = obj.precio_final
        if pf < obj.precio:
            return format_html('<span style="color:#e53e3e;font-weight:bold;">${:,.0f}</span> <small style="text-decoration:line-through;color:#999;">${:,.0f}</small>', pf, obj.precio)
        return format_html('${:,.0f}', obj.precio)

    @admin.display(description='Stock')
    def indicador_stock(self, obj):
        if obj.stock == 0:
            color, icono = '#e53e3e', '🔴'
        elif obj.stock <= obj.stock_minimo:
            color, icono = '#dd6b20', '🟡'
        else:
            color, icono = '#38a169', '🟢'
        return format_html(
            '<span style="color:{};font-weight:bold;">{} {}</span>',
            color, icono, obj.stock
        )


@admin.register(Canasto)
class CanastoAdmin(admin.ModelAdmin):
    list_display = ('nombre', 'tipo', 'precio_base', 'activo')
    list_filter = ('tipo', 'activo')


@admin.register(Pedido)
class PedidoAdmin(admin.ModelAdmin):
    list_display = ('id', 'nombre_cliente', 'ciudad', 'total', 'estado', 'metodo_pago', 'creado')
    list_filter = ('estado', 'metodo_pago', 'ciudad')
    search_fields = ('nombre_cliente', 'email')
    readonly_fields = ('creado', 'stripe_payment_id')
    inlines = [ItemPedidoInline]


@admin.register(Suscripcion)
class SuscripcionAdmin(admin.ModelAdmin):
    list_display = ('nombre_cliente', 'email', 'telefono', 'ciudad', 'frecuencia', 'tiene_canasto', 'tiene_tarro_organicos', 'proxima_entrega', 'activa')
    list_filter = ('frecuencia', 'activa', 'tiene_canasto', 'tiene_tarro_organicos', 'ciudad')
    search_fields = ('nombre_cliente', 'email', 'telefono')
    list_editable = ('activa', 'tiene_canasto', 'tiene_tarro_organicos', 'proxima_entrega')
    readonly_fields = ('creada',)
    fieldsets = (
        ('Cliente', {
            'fields': ('nombre_cliente', 'email', 'telefono', 'ciudad', 'direccion')
        }),
        ('Suscripción', {
            'fields': ('frecuencia', 'activa', 'proxima_entrega', 'ultimo_pedido', 'notas')
        }),
        ('Economía circular', {
            'fields': ('tiene_canasto', 'tiene_tarro_organicos'),
            'description': 'Registro del material físico que tiene el cliente en su poder'
        }),
        ('Sistema', {
            'fields': ('creada',),
            'classes': ('collapse',)
        }),
    )


class IngredienteRecetaInline(admin.TabularInline):
    model = IngredienteReceta
    extra = 1


@admin.register(Receta)
class RecetaAdmin(admin.ModelAdmin):
    list_display = ('nombre', 'temporada_display', 'disponible_ahora', 'creada')
    list_filter = ('creada',)
    search_fields = ('nombre', 'descripcion')
    inlines = [IngredienteRecetaInline]
    readonly_fields = ('disponible_ahora',)

    @admin.display(description='Temporada', boolean=False)
    def temporada_display(self, obj):
        meses = {1:'Ene',2:'Feb',3:'Mar',4:'Abr',5:'May',6:'Jun',
                 7:'Jul',8:'Ago',9:'Sep',10:'Oct',11:'Nov',12:'Dic'}
        return ', '.join(meses.get(m, str(m)) for m in (obj.meses_temporada or []))

    @admin.display(description='Disponible ahora', boolean=True)
    def disponible_ahora(self, obj):
        return obj.disponible_ahora


class ItemPaqueteInline(admin.TabularInline):
    model = ItemPaquete
    extra = 1


@admin.register(Paquete)
class PaqueteAdmin(admin.ModelAdmin):
    list_display = ('nombre', 'tipo', 'mercado', 'activo', 'creado')
    list_filter = ('tipo', 'activo')
    search_fields = ('nombre', 'descripcion')
    inlines = [ItemPaqueteInline]


@admin.register(ClienteCorporativo)
class ClienteCorporativoAdmin(admin.ModelAdmin):
    list_display = ('nombre_empresa', 'mercado', 'contacto', 'descuento_porcentaje', 'limite_credito', 'activo')
    list_filter = ('mercado', 'activo')
    search_fields = ('nombre_empresa', 'nit')


@admin.register(DocumentoRAG)
class DocumentoRAGAdmin(admin.ModelAdmin):
    list_display = ('titulo', 'tipo', 'mercado', 'activo', 'actualizado')
    list_filter = ('tipo', 'mercado', 'activo')
    search_fields = ('titulo', 'contenido')
    readonly_fields = ('actualizado',)


class MensajeAgenteInline(admin.TabularInline):
    model = MensajeAgente
    extra = 0
    readonly_fields = ('rol', 'contenido', 'creado')
    can_delete = False


@admin.register(ConversacionAgente)
class ConversacionAgenteAdmin(admin.ModelAdmin):
    list_display = ('id', 'usuario', 'mercado', 'session_id', 'creada')
    list_filter = ('mercado',)
    search_fields = ('usuario__username', 'usuario__email', 'session_id')
    readonly_fields = ('creada',)
    inlines = [MensajeAgenteInline]
